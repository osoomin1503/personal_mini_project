#2-4.
#엑셀 모듈 import
import openpyxl

# PythonLecture이라는 class 정의
class PythonLecture:
    # 생성자 함수 정의
    def __init__(self, score1, score2, score3, absent):
        # 사용자의 입력값을 self.입력값의 형태로 지정
        self.score1 = score1
        self.score2 = score2
        self.score3 = score3
        self.absent = absent
    # 3개의 시험의 평균 점수를 산출하는 함수 정의
    def average_score(self):
        average = (self.score1 + self.score2 + self.score3) / 3
        return average
    # 3개의 점수를 20,30,50% 가중치로 평균을 산출하는 함수 정의
    def weigh_average(self):
        average = 0.2 * self.score1 + 0.3 * self.score2 + 0.5 * self.score3
        return average
    # 결석 횟수가 5회 이상이면 성적과 상관없이 F학점을 산출하고,
    # weigh_average 함수의 반환값을 최종성적으로 받아 학점으로 산출하는 함수 정의
    def grade(self, avg_score):
        # 결석 횟수가 5회 이상일 때
        if self.absent >= 5:
            return 'F'

        # avg_score이 각 구간에 속할 때 반환되는 학점
        if 100 >= avg_score >= 90:
            return 'A'
        elif 80 <= avg_score < 90:
            return 'B'
        elif 70 <= avg_score < 80:
            return 'C'
        elif 60 <= avg_score < 70:
            return 'D'
        elif avg_score < 60:
            return 'F'

# 엑셀 모듈을 여는 함수 실행
wb_obj = openpyxl.load_workbook("score_table.xlsx")
# 현재 시트에 대한 객체를 엑셀 파일을 활성화시킴으로서 설정
sheet_obj = wb_obj.active
# final 엑셀 파일에 추가적으로 구현되어야할 문구를 행과 열의 인덱스를 이용해 추가.
sheet_obj.cell(1, 6, "Final Score")
sheet_obj.cell(1, 7, "Final Grade")

# 엑셀 파일에 각 cell의 값을 대입하기 위한 반복문
for i in range(2, 12):
    # 원하는 cell의 위치를 행과 열의 인덱스로 지정하고 value라는 변수에 저장
    score1 = sheet_obj.cell(row=i, column=2).value
    score2 = sheet_obj.cell(row=i, column=3).value
    score3 = sheet_obj.cell(row=i, column=4).value
    absent = sheet_obj.cell(row=i, column=5).value
    # 변수값을 class에 입력
    Student = PythonLecture(score1, score2, score3, absent)

    # class의 weigh_average 함수로 평균을 구하고 소수점이 있는 숫자로 구현.
    avg_score = float(Student.weigh_average())
    # class의 grade 함수로 위의 avg_score에 대한 학점을 구하도록 설정
    grade = Student.grade(avg_score)

    # 각 행과 6번째 열에 구한 평균값을 넣고, 7번째 열에는 학점을 넣도록 구현
    sheet_obj.cell(i, 6, avg_score)
    sheet_obj.cell(i, 7, grade)

# score_table_final.xlsx로 위의 내용을 저장
wb_obj.save("score_table_final.xlsx")
# 파일 닫기
wb_obj.close()