#final
import openpyxl
import time

class BankAccount:
    def __init__(self, name='None', balance=0):
        self.name = name
        self.balance = balance
        print(name, '님 환영합니다.')
        print('초기 금액 ', balance, '원으로 계좌가 만들어졌습니다.')

    def withdraw(self, amount):
        try:
            if amount > self.balance:
                raise AccountBalanceException
            else:
                self.balance -= amount
                print('통장에 ', amount, '원이 출금되었음')
        except AccountBalanceException as e:
            print(e)
        finally:
            print('현재 잔액은 ', self.balance, '원입니다')

    def deposit(self, amount):
        try:
            if amount < 0:
                raise InvalidTransactionException
            else:
                self.balance += amount
                print('통장에 ', amount, '원이 입금되었음')
        except InvalidTransactionException as e:
            print(e)
        finally:
            print('현재 잔액은 ', self.balance, '입니다')

class AccountException(Exception):
    pass

class AccountBalanceException(AccountException):
    def __str__(self):
        return 'Account Balance Exception Occurs: Check your balance'

class InvalidTransactionException(AccountException):
    def __str__(self):
        return 'Invalid Transaction Exception Occurs: Check your amount'

class Atm(BankAccount) :

    def __init__(self, name, balance=0):
        self.saveRowCnt = 2         # 엑셀 위치 저장 (신규 등록 고객은 맨뒤)
        self.OKPassword = False     # 비밀번호 입력이 맞는지 확인한 결과 저장

        # 등록된 고객인지 확인
        if self.isRegisteredUser(name) == True:
            # 등록고객이면 비밀번호를 확인한 후 거래를 시작한다.
            input_password = self.inputPassword()
            self.checkPassword(input_password)
        else:
            # 등록고객이 아니면 고객 정보를 입력받아 새롭게 등록한다.
            input_password = self.inputPassword()
            if input_password > 0 :
                self.OKPassword = True
                input_balance = int(input("초기 입금액을 입력하세요 : "))
                super().__init__(name, input_balance)
                self.saveNewUser(name, input_password, input_balance)

    #--------------------------------------------------------------------------
    # 등록된 고객인지 확인 : UserList.xlsx을 읽어 동일한 이름이 있는지 확인한다.
    # 거래할 고객 정보가 저장된 엑셀의 row 위치를 saveRowCnt에 저장한다.
    #--------------------------------------------------------------------------
    def isRegisteredUser(self, userName):
        wb_obj = openpyxl.load_workbook("UserList.xlsx")
        sheet_obj = wb_obj.active
        i = 2
        while True:
            name = sheet_obj.cell(row=i, column=1).value
            password = sheet_obj.cell(row=i, column=2).value
            balance = sheet_obj.cell(row=i, column=3).value

            if name == userName:
                self.name = name
                self.password = int(password)
                self.balance = int(balance)
                wb_obj.close()
                return True
            elif name == None:
                break
            else:
                i = i + 1
                self.saveRowCnt = i
        wb_obj.close()
        return False

    #--------------------------------------------------------------------------
    # 등록된 고객이 아니면 입력된 사용자 정보를 새롭게 등록한다.
    #--------------------------------------------------------------------------
    def saveNewUser(self, name, password, balance):
        wb_obj = openpyxl.load_workbook("UserList.xlsx")
        sheet_obj = wb_obj.active
        sheet_obj.cell(self.saveRowCnt, 1, name)
        sheet_obj.cell(self.saveRowCnt, 2, password)
        sheet_obj.cell(self.saveRowCnt, 3, balance)
        wb_obj.save("UserList.xlsx")
        wb_obj.close()

    #--------------------------------------------------------------------------
    # 비밀번호를 4자리 이하로 입력하면 재입력을 받는다.
    #--------------------------------------------------------------------------
    def inputPassword(self):
        password = int(input("비밀번호 4자리를 입력하세요 : "))
        if 1000 <= password < 10000:
            return password
        return 0

    #--------------------------------------------------------------------------
    # 비밀번호가 맞는지 확인 후 거래한다.
    #--------------------------------------------------------------------------
    def checkPassword(self, input_password):
        if input_password == self.password:
            print(self.name, '님은 등록된 고객입니다.')
            self.OKPassword = True
        else :
            print(self.name, '님! 비밀번호 입력 오류입니다. 처음부터 다시 시작하세요.')
            self.OKPassword = False

    #--------------------------------------------------------------------------
    # 엑셀파일에 거래 후 잔액을 저장한다.
    #--------------------------------------------------------------------------
    def saveBalance(self):
        wb_obj = openpyxl.load_workbook("UserList.xlsx")
        sheet_obj = wb_obj.active
        sheet_obj.cell(self.saveRowCnt, 3, self.balance)
        wb_obj.save("UserList.xlsx")
        wb_obj.close()

    #--------------------------------------------------------------------------
    # 거래명세표를 출력한다.
    #--------------------------------------------------------------------------
    def printReport(self, name, choice, amount ) :
        print('-' * 30)
        print('거래 고객 :', name)

        dateTime = time.strftime('%Y/%m/%d',time.localtime(time.time())) \
                   + time.strftime(' %H:%M:%S',time.localtime(time.time()))
        print('거래 시간 :', dateTime )
        if choice == 1 :
            print('거래 구분 : 입금')
        else :
            print('거래 구분 : 출금')
        print('거래 금액 :', amount, '원')
        print('거래후 잔액 :', self.balance, '원')
        print('-' * 30)

while True :
    print('=' * 30)
    print('Soomin Bank ATM')
    print('=' * 30)
    atm_name = input("거래를 원하시면 고객님 이름을 입력하세요 : ")
    atm_bank = Atm(atm_name)

    # 비밀번호 입력이 맞지 않으면 거래를 중단한다.
    if atm_bank.OKPassword == False :
        continue

    while True :
        print('=' * 40)
        print("1. 입금, 2. 출금, 3. 잔액조회, 4. 종료 ")
        print('=' * 40)
        choice = int(input("거래 번호를 선택하세요 : "))

        # 잔액조회
        if choice == 3:
            print('현재 잔액은 ', atm_bank.balance, '원입니다')
            continue
        # 종료
        elif choice == 4 :
            print('거래를 종료합니다')
            break

        amount = 0
        balance = atm_bank.balance
        # 입금
        if choice == 1 :
            amount = int(input("입금할 금액을 입력하세요 : "))
            atm_bank.deposit(amount)
        #출금
        elif choice == 2 :
            amount = int(input("출금할 금액을 입력하세요 : "))
            atm_bank.withdraw(amount)

        # 거래 후 잔액 변화가 있을 때만, 잔액을 엑셀에 저장하고 명세표 출력 여부를 노출한다.
        if balance != atm_bank.balance :
            atm_bank.saveBalance()
            isReport = str(input("명세표를 출력하시겠습니까?(Y/N) "))
            if isReport == 'Y' or isReport == 'y' :
                atm_bank.printReport(atm_name, choice, amount)
            elif isReport == 'N' or isReport == 'n' :
                pass

        isRepeat = str(input("거래를 계속하시겠습니까?(Y/N) "))
        if isRepeat == 'N' or isRepeat =='n' :
            print('안녕히 가십시오.')
            break
        elif isRepeat == 'Y' or isRepeat == 'y' :
            pass
